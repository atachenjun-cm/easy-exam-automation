import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from score_feedback_exporter import export_score_feedback


class ScoreFeedbackExporterTest(unittest.TestCase):
    def test_exports_scores_from_template_with_text_identifiers(self):
        root = Path(__file__).resolve().parents[1]
        template = root / "template" / "成绩单模板.xlsx"
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            payload = temp / "payload.json"
            output = temp / "score.xlsx"
            rows = []
            for index in range(6):
                rows.append(
                    {
                        "name": f"考生{index + 1}",
                        "gender": "男" if index % 2 == 0 else "女",
                        "identity_id": f"00123456789012345{index}",
                        "mobile": f"0138000000{index:02d}",
                        "email": "" if index == 0 else f"user{index}@example.com",
                        "course": "综合能力",
                        "permit": f"000{index + 1}",
                        "exam_status": "已完成" if index == 0 else ("未开考" if index == 1 else "异常状态"),
                        "score": "" if index == 1 else 80 + index,
                        "violation": "" if index != 2 else "作弊",
                    }
                )
            payload.write_text(
                json.dumps(
                    {
                        "examName": "客户招聘笔试",
                        "examTime": "2026-07-25 09:00 ~ 2026-07-25 10:30",
                        "processedDate": "2026年6月29日",
                        "rows": rows,
                    },
                    ensure_ascii=False,
                ),
                "utf-8",
            )

            result = export_score_feedback(template, payload, output)

            self.assertTrue(result["ok"])
            workbook = load_workbook(output)
            sheet = workbook.active
            self.assertEqual(sheet["B4"].value, "客户招聘笔试")
            self.assertEqual(sheet["C4"].value, "2026-07-25 09:00 ~ 2026-07-25 10:30")
            self.assertEqual(sheet["A7"].value, "考生1")
            self.assertEqual(sheet["H7"].value, "参考")
            self.assertEqual(sheet["H8"].value, "缺考")
            self.assertEqual(sheet["J7"].value, "无")
            self.assertEqual(sheet["C7"].number_format, "@")
            self.assertEqual(sheet["D7"].number_format, "@")
            self.assertEqual(sheet["G7"].number_format, "@")
            self.assertEqual(sheet["C7"].value, "001234567890123450")
            self.assertEqual(sheet["D7"].value, "013800000000")
            footer_rows = [
                row[0].row
                for row in sheet.iter_rows()
                for cell in row
                if cell.value == "全美在线（北京）科技股份有限公司"
            ]
            self.assertTrue(footer_rows)
            self.assertGreater(footer_rows[0], 11)
            self.assertEqual(sheet.cell(footer_rows[0] + 1, 6).value, "2026年6月29日")

    def test_exports_valid_without_score_as_absent_with_placeholder(self):
        root = Path(__file__).resolve().parents[1]
        template = root / "template" / "成绩单模板.xlsx"
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            payload = temp / "payload.json"
            output = temp / "score.xlsx"
            payload.write_text(
                json.dumps(
                    {
                        "examName": "缺考状态测试",
                        "examTime": "2026-06-25 19:00 ~ 2026-06-25 21:00",
                        "rows": [
                            {
                                "name": "李科",
                                "permit": "13208164907",
                                "exam_status": "valid",
                                "score": "",
                                "violation": "",
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
                "utf-8",
            )

            result = export_score_feedback(template, payload, output)

            self.assertTrue(result["ok"])
            sheet = load_workbook(output).active
            self.assertEqual(sheet["H7"].value, "缺考")
            self.assertEqual(sheet["I7"].value, "--")


if __name__ == "__main__":
    unittest.main()
