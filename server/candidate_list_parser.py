#!/usr/bin/env python3
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

FULL_NAME_ALIASES = {"姓名", "考生姓名", "full_name", "name"}
IDENTITY_ID_ALIASES = {"证件号", "身份证号", "身份证", "证件号码", "identity_id", "ssn"}
PERMIT_ALIASES = {"准考证号", "考号", "考生编号", "permit"}
COURSE_CODE_ALIASES = {"科目编号", "易考科目编号", "course_code"}
MOBILE_ALIASES = {"手机号码", "手机号", "手机", "联系电话", "电话", "mobile", "phone"}
EMAIL_ALIASES = {"邮箱", "邮箱地址", "电子邮箱", "电子邮件", "邮件", "email", "mail"}
REQUIRED_FIELDS = ("permit", "full_name")
TARGET_FIELDS = ("permit", "full_name", "identity_id", "course_code", "mobile", "email")
TEMPLATE_HEADERS = ("准考证号", "姓名", "身份证号", "科目编号", "手机号码", "邮箱")
DISALLOWED_CUSTOM_FIELD_NAMES = {"姓名", "身份证号", "证件号", "准考证号", "科目编号", "科目名称"}
MAX_CUSTOM_FIELDS = 30
SCIENTIFIC_RE = re.compile(r"^\s*\d+(?:\.\d+)?[eE]\+?\d+\s*$")


def normalize_header(value):
    return str(value or "").strip().replace(" ", "").replace("\u3000", "")


def cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def detect_mapping(columns):
    normalized = {normalize_header(col).lower(): col for col in columns}

    def pick(aliases):
        for alias in aliases:
            found = normalized.get(normalize_header(alias).lower())
            if found:
                return found
        return ""

    return {
        "full_name": pick(FULL_NAME_ALIASES),
        "identity_id": pick(IDENTITY_ID_ALIASES),
        "permit": pick(PERMIT_ALIASES),
        "course_code": pick(COURSE_CODE_ALIASES),
        "mobile": pick(MOBILE_ALIASES),
        "email": pick(EMAIL_ALIASES),
    }


def canonical_import_field_name(value):
    name = str(value or "").strip()
    normalized = normalize_header(name).lower()
    if normalized in {normalize_header(alias).lower() for alias in MOBILE_ALIASES}:
        return "手机号码"
    if normalized in {normalize_header(alias).lower() for alias in EMAIL_ALIASES}:
        return "邮箱"
    return name


def custom_field_candidates(columns, mapping):
    return [
        column
        for column in columns
        if str(column or "").strip()
        and str(column or "").strip() not in DISALLOWED_CUSTOM_FIELD_NAMES
        and canonical_import_field_name(column) not in {"手机号码", "邮箱"}
    ]


def validate_custom_fields(custom_fields):
    errors = []
    normalized = []
    seen = set()
    enabled_fields = [field for field in custom_fields or [] if field.get("enabled")]
    if len(enabled_fields) > MAX_CUSTOM_FIELDS:
        errors.append(f"自定义字段最多支持 {MAX_CUSTOM_FIELDS} 个")
    for index, field in enumerate(enabled_fields, start=1):
        source_column = str(field.get("source_column") or "").strip()
        target_name = canonical_import_field_name(field.get("target_name") or "")
        if not target_name:
            errors.append(f"第 {index} 个自定义字段名称不能为空")
            continue
        if target_name in DISALLOWED_CUSTOM_FIELD_NAMES:
            errors.append(f"自定义字段不能作为导入信息项：{target_name}")
            continue
        if target_name in seen:
            errors.append(f"自定义字段名称重复：{target_name}")
            continue
        seen.add(target_name)
        normalized.append({"source_column": source_column, "target_name": target_name, "enabled": True})
    return errors, normalized


def read_csv(path):
    last_error = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with open(path, "r", encoding=encoding, newline="") as fh:
                reader = csv.reader(fh)
                rows = list(reader)
            return rows
        except UnicodeDecodeError as exc:
            last_error = exc
    raise RuntimeError(f"CSV 编码无法识别：{last_error}")


def read_xlsx(path):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("当前 Python 环境缺少 openpyxl，请先执行 python3 -m pip install -r requirements.txt") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([cell_to_text(cell) for cell in row])
    return rows


def read_xls(path):
    try:
        import xlrd
    except Exception as exc:
        raise RuntimeError("当前 Python 环境缺少 xlrd，无法解析 .xls；请转为 .xlsx 或安装 xlrd") from exc

    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        rows.append([cell_to_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)])
    return rows


def read_table(path):
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix == ".xlsx":
        return read_xlsx(path)
    if suffix == ".xls":
        return read_xls(path)
    raise RuntimeError("文件格式不支持，仅支持 .xlsx、.xls、.csv")


def trim_empty_tail(values):
    result = list(values)
    while result and not str(result[-1] or "").strip():
        result.pop()
    return result


def normalize_rows(rows):
    rows = [trim_empty_tail(row) for row in rows if any(str(cell or "").strip() for cell in row)]
    if not rows:
        raise RuntimeError("名单为空")
    headers = [str(cell or "").strip() for cell in rows[0]]
    if not any(headers):
        raise RuntimeError("未识别到表头")
    data = []
    for index, row in enumerate(rows[1:], start=2):
        item = {"__row": index}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            item[header] = cell_to_text(row[col_index]) if col_index < len(row) else ""
        data.append(item)
    return headers, data


def build_candidates(raw_rows, mapping, custom_fields=None):
    custom_errors, normalized_custom_fields = validate_custom_fields(custom_fields or [])
    if custom_errors:
        raise ValueError("；".join(custom_errors))
    candidates = []
    for row in raw_rows:
        custom_values = {}
        for field in normalized_custom_fields:
            target_name = field["target_name"]
            source_column = field["source_column"]
            custom_values[target_name] = cell_to_text(row.get(source_column, "")) if source_column else ""
        candidates.append(
            {
                "__row": row.get("__row"),
                "permit": cell_to_text(row.get(mapping.get("permit", ""), "")),
                "full_name": cell_to_text(row.get(mapping.get("full_name", ""), "")),
                "identity_id": cell_to_text(row.get(mapping.get("identity_id", ""), "")),
                "course_code": cell_to_text(row.get(mapping.get("course_code", ""), "")),
                "mobile": cell_to_text(row.get(mapping.get("mobile", ""), "")),
                "email": cell_to_text(row.get(mapping.get("email", ""), "")),
                "custom_fields": custom_values,
            }
        )
    return candidates


def validate_candidates(candidates, mapping):
    errors = []
    warnings = []
    for field in REQUIRED_FIELDS:
        if not mapping.get(field):
            errors.append(f"缺少字段映射：{field}")

    duplicate_maps = {"permit": defaultdict(list), "identity_id": defaultdict(list)}
    for idx, row in enumerate(candidates, start=2):
        row_num = row.get("__row") or idx
        for field in REQUIRED_FIELDS:
            value = cell_to_text(row.get(field))
            if not value:
                errors.append(f"第 {row_num} 行缺少 {field}")
        for field in ("permit", "identity_id", "mobile"):
            value = cell_to_text(row.get(field))
            if value and SCIENTIFIC_RE.match(value):
                errors.append(f"第 {row_num} 行 {field} 为科学计数法格式，请修正原始文件后再导入")
        for field, bucket in duplicate_maps.items():
            value = cell_to_text(row.get(field))
            if value:
                bucket[value].append(row_num)

    for value, row_nums in duplicate_maps["permit"].items():
        if len(row_nums) > 1:
            errors.append(f"准考证号重复：{value}，行号：{','.join(map(str, row_nums))}")
    for value, row_nums in duplicate_maps["identity_id"].items():
        if len(row_nums) > 1:
            errors.append(f"证件号重复：{value}，行号：{','.join(map(str, row_nums))}")

    return errors, warnings


def parse(path):
    rows = read_table(path)
    columns, raw_rows = normalize_rows(rows)
    mapping = detect_mapping(columns)
    candidates = build_candidates(raw_rows, mapping)
    errors, warnings = validate_candidates(candidates, mapping)
    custom_candidates = custom_field_candidates(columns, mapping)
    return {
        "columns": columns,
        "headers": columns,
        "mapping": mapping,
        "auto_mapping": mapping,
        "custom_field_candidates": custom_candidates,
        "rawRows": raw_rows,
        "candidates": [{k: v for k, v in row.items() if k in (*TARGET_FIELDS, "custom_fields")} for row in candidates],
        "preview": [{k: row.get(k, "") for k in TARGET_FIELDS} for row in candidates[:20]],
        "preview_rows": raw_rows[:20],
        "totalCount": len(candidates),
        "errors": errors,
        "warnings": warnings,
    }


def write_template(payload_path, output_path):
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("当前 Python 环境缺少 openpyxl，请先执行 python3 -m pip install -r requirements.txt") from exc

    payload = json.loads(Path(payload_path).read_text("utf-8"))
    candidates = payload.get("candidates") or []
    errors, warnings = validate_candidates(candidates, {field: field for field in TARGET_FIELDS})
    if errors:
        return {"ok": False, "errors": errors, "warnings": warnings}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "candidates"
    sheet.append(list(TEMPLATE_HEADERS))
    for row in candidates:
        sheet.append([cell_to_text(row.get(field)) for field in TARGET_FIELDS])

    for col in ("A", "B", "C", "D", "E", "F"):
        for cell in sheet[col]:
            cell.number_format = "@"

    workbook.save(output_path)
    return {"ok": True, "path": str(output_path), "errors": [], "warnings": warnings}


def main():
    if len(sys.argv) < 3:
        raise RuntimeError("Usage: candidate_list_parser.py parse <path> | template <payload.json> <output.xlsx>")
    command = sys.argv[1]
    if command == "parse":
        result = parse(sys.argv[2])
    elif command == "template":
        if len(sys.argv) < 4:
            raise RuntimeError("template command requires output path")
        result = write_template(sys.argv[2], sys.argv[3])
    else:
        raise RuntimeError(f"Unknown command: {command}")
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        sys.exit(1)
