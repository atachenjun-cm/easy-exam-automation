#!/usr/bin/env python3

import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


BOOL_TRUE = {"是", "需要", "开启", "开启录制", "true", "yes", "y", "1"}
BOOL_FALSE = {"否", "不需要", "关闭", "false", "no", "n", "0"}


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_bool(value, default=False):
    text = normalize_text(value).lower()
    if not text:
        return default
    if text in BOOL_TRUE:
        return True
    if text in BOOL_FALSE:
        return False
    return default


def parse_enabled_text(value, default=False):
    text = normalize_text(value)
    if not text:
        return default
    lowered = text.lower()
    if lowered in BOOL_FALSE or any(token in lowered for token in ("不需要", "无需", "不开", "关闭", "否")):
        return False
    return True


def parse_minutes(value):
    text = normalize_text(value)
    if not text:
        return None
    match = re.search(r"(\d+)", text)
    return int(match.group(1)) if match else None


def normalize_exam_address(value):
    text = normalize_text(value)
    if "统一" in text:
        return "统一考试地址"
    if "独立" in text:
        return "独立考试地址"
    return text


def split_subject_names(value):
    text = normalize_text(value)
    if not text:
        return []
    parts = re.split(r"[\n,，、;；]+", text)
    return [part.strip() for part in parts if part and part.strip()]


def parse_subjects_text(value):
    return split_subject_names(value)


def split_form_codes(value):
    text = normalize_text(value)
    if not text:
        return []
    parts = re.split(r"[\n,，;；\s]+", text)
    return [part.strip() for part in parts if part and part.strip()]


def normalize_course_record(name="", code="", form_codes=""):
    course_name = normalize_text(name)
    course_code = normalize_text(code)
    forms = split_form_codes(form_codes)
    if not course_name and not course_code:
        return None
    return {
        "name": course_name,
        "code": course_code,
        "form_codes": forms,
    }


def parse_course_records_text(value):
    text = normalize_text(value)
    if not text:
        return []
    records = []
    for line in re.split(r"[\n;；]+", text):
        parts = [part.strip() for part in re.split(r"[|｜\t]", line) if part and part.strip()]
        if len(parts) >= 2:
            record = normalize_course_record(
                parts[0],
                parts[1],
                parts[2] if len(parts) > 2 else "",
            )
            if record and record["name"] and record["code"]:
                records.append(record)
    return records


def parse_time_range(value):
    text = normalize_text(value)
    if not text:
        return None, None
    parts = [part.strip() for part in re.split(r"\s*(?:-|–|—|~|至)\s*", text) if part.strip()]
    if len(parts) != 2:
        return None, None
    return parse_datetime(parts[0]), parse_datetime(parts[1])


def parse_datetime(value):
    text = normalize_text(value)
    if not text:
        return None
    formats = [
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def format_datetime(dt):
    return dt.strftime("%Y/%m/%d %H:%M") if dt else ""


def slugify_subjects(subjects):
    safe = [re.sub(r"[^\w\u4e00-\u9fff-]+", "", name) for name in subjects]
    joined = "".join(filter(None, safe)) or "科目"
    return f"科目导入_{joined[:40]}.xlsx"


def build_subject_workbook(subjects, output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / slugify_subjects(subjects)
    wb = Workbook()
    ws = wb.active
    ws.title = "科目信息"
    ws.append(["序号", "科目"])
    for index, subject in enumerate(subjects, start=1):
        ws.append([index, subject])
    wb.save(path)
    return str(path)


def read_subjects(sheet):
    subjects = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        name = normalize_text(row[1] if len(row) > 1 else "")
        if name:
            subjects.extend(split_subject_names(name))
    return subjects


def read_course_records(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_text(value).lower() for value in rows[0]]

    def find_index(*names):
        normalized_names = [normalize_text(name).lower() for name in names]
        for index, header in enumerate(headers):
            if header in normalized_names:
                return index
        for index, header in enumerate(headers):
            if any(name and name in header for name in normalized_names):
                return index
        return None

    name_index = find_index("name", "科目名称", "科目")
    code_index = find_index("code", "course_code", "科目编号", "科目代码")
    form_index = find_index("form_codes", "form code", "试卷编码", "试卷编号", "试卷代码")

    if name_index is None:
        name_index = 1
    if code_index is None:
        code_index = 2
    if form_index is None:
        form_index = 3

    records = []
    for row in rows[1:]:
        if not row:
            continue
        name = row[name_index] if len(row) > name_index else ""
        code = row[code_index] if len(row) > code_index else ""
        form_codes = row[form_index] if len(row) > form_index else ""
        if normalize_text(name) and not normalize_text(code) and not normalize_text(form_codes):
            for subject in split_subject_names(name):
                record = normalize_course_record(subject, "", "")
                if record:
                    records.append(record)
            continue
        record = normalize_course_record(name, code, form_codes)
        if record and (record["name"] or record["code"]):
            records.append(record)
    return records


def build_generated_courses(subjects, start_dt):
    if not subjects or not start_dt:
        return []
    prefix = start_dt.strftime("%Y%m%d")
    courses = []
    for index, subject in enumerate(subjects, start=1):
        course_code = f"{prefix}-01-{index:02d}"
        courses.append(
            {
                "name": subject,
                "code": course_code,
                "form_codes": [course_code],
            }
        )
    return courses


def build_preview(config):
    preview = []

    def add(stage, item, value, action):
        preview.append([stage, item, value or "空", action])

    add("基础信息", "考试名称", config["examName"], "自动填写")
    add(
        "基础信息",
        "考试时间",
        f'{config["startTimeDisplay"]} 至 {config["endTimeDisplay"]}',
        "人工核对",
    )
    add("基础信息", "提前登录时间", f'{config["earlyLoginMinutes"] or 0} 分钟', "自动填写")
    add("基础信息", "限制迟到时间", f'{config["lateLimitMinutes"] or 0} 分钟', "自动填写")
    add("基础信息", "试卷扣时规则", config["timeRule"] or "系统默认", "自动选择")
    add("基础信息", "考试地址", config["examAddress"] or "系统默认", "按需求单选择")
    add("基础信息", "考前等待提示", config["preLoginPrompt"] or "空", "自动填写")
    add("基础信息", "欢迎语", config["welcomeText"] or "空", "自动填写")
    add("科目管理", "批量导入科目", "、".join(config["subjects"]) or "空", "下载后台模板后导入")
    add("选择试卷", "跳过试卷设置", "是", "自动跳过")
    add("个人信息", "考生可见字段", "姓名, 身份证号", "自动勾选")
    add("个人信息", "允许编辑字段", "无", "自动取消")
    add("个人信息", "必填字段", "无", "自动取消")
    add("开考前", "考试承诺书", config["pledgeContent"] or "否", "自动填写")
    add(
        "考试中",
        "视频监控/录制",
        f'{"是" if config["videoMonitor"] else "否"} / {"是" if config["videoRecord"] else "否"}',
        "自动勾选",
    )
    if config["videoMonitor"]:
        add("考试中", "登录验证", "自动验证；考后公安验证", "视频监控默认启用")
        add("考试中", "作弊侦测", "基础版AI", "视频监控默认启用")
    add("考试中", "鹰眼监控", "是" if config["hawkeye"] else "否", "自动勾选")
    add(
        "考试中",
        "锁定考试",
        (
            f'客户端考试；登录限制 {config["clientLoginLimit"]} 次'
            if config["clientExam"]
            else f'网页考试；允许离开 {config["leaveLimit"]} 次'
            if config["webExam"] and config["leaveLimit"] is not None
            else "否"
        ),
        "自动勾选",
    )
    add("考试中", "答题水印", "是" if config["watermark"] else "否", "自动勾选")
    add("考试中", "禁止复制", "是" if config["disableCopy"] else "否", "自动勾选")
    add("考试后", "人工判分", config["manualScoreText"] or ("是" if config["manualScore"] else "否"), "按需求单配置")
    if config["mockExamEnabled"]:
        add("试考", "试考名称", config["mockExamName"], "自动新建")
        add(
            "试考",
            "试考时间",
            f'{config["mockStartTimeDisplay"]} 至 {config["mockEndTimeDisplay"]}',
            "自动填写",
        )
        add("试考", "提前登录时间", "不设置", "自动跳过")
        add("试考", "限制迟到时间", "不设置", "自动跳过")
        add("试考", "试卷扣时规则", "不扣时", "自动选择")
        add("试考", "科目设置", "跳过", "自动跳过")
    add("完成", "最终创建", "点击创建完成", "自动创建")
    return preview


def parse_workbook(path_str):
    path = Path(path_str)
    wb = load_workbook(path, data_only=True)
    sheet = wb["业务需求单"]
    subjects_sheet = wb["科目信息"] if "科目信息" in wb.sheetnames else None

    field_map = {}
    for row in sheet.iter_rows(min_row=5, values_only=True):
        if len(row) < 4:
            continue
        item = normalize_text(row[2])
        value = normalize_text(row[3])
        if item:
            field_map[item] = value

    def get_field(*names):
        for name in names:
            if name in field_map and field_map[name]:
                return field_map[name]
        for key, value in field_map.items():
            if value and any(name in key for name in names):
                return value
        return ""

    start_dt, end_dt = parse_time_range(get_field("考试日期时间", "考试时间", "考试起止时间"))
    mock_start_dt, mock_end_dt = parse_time_range(get_field("试考日期时间", "试考时间", "试考起止时间"))
    courses = parse_course_records_text(get_field("科目信息", "科目"))
    subjects = [course["name"] or course["code"] for course in courses]
    if not subjects:
        subjects = parse_subjects_text(get_field("科目信息", "科目"))
    if subjects_sheet:
        sheet_courses = read_course_records(subjects_sheet)
        if sheet_courses:
            courses = sheet_courses
            subjects = [course["name"] or course["code"] for course in courses]
        elif not subjects:
            subjects = read_subjects(subjects_sheet)
    if subjects and not [course for course in courses if course.get("code")]:
        courses = build_generated_courses(subjects, start_dt)
    subject_import_dir = path.parent / "exam_request"
    subject_import_path = build_subject_workbook(subjects, subject_import_dir) if subjects else ""

    time_rule = get_field("试卷扣时规则", "扣时规则", "扣时")
    exam_type = get_field("考试类型")
    exam_address = normalize_exam_address(get_field("考试地址", "考试网址类型", "考试链接类型"))
    manual_score_text = get_field("人工判分", "主观题判分", "判分方式")

    config = {
        "examName": get_field("考试名称", "考试名"),
        "u8Code": get_field("U8代码", "U8编码", "U8", "项目代码"),
        "projectManager": get_field("项目经理", "项目负责人", "负责人"),
        "customerName": get_field("客户名称", "客户", "单位名称", "委托单位"),
        "candidateCount": parse_minutes(get_field("人次", "考生人数", "人数")) or "",
        "startTimeDisplay": format_datetime(start_dt),
        "endTimeDisplay": format_datetime(end_dt),
        "startTimeIso": start_dt.isoformat() if start_dt else "",
        "endTimeIso": end_dt.isoformat() if end_dt else "",
        "earlyLoginMinutes": parse_minutes(get_field("提前登录时间", "提前登录分钟", "提前登录")),
        "lateLimitMinutes": parse_minutes(get_field("限制迟到时间", "限制迟到分钟", "限制迟到")),
        "timeRule": time_rule,
        "examAddress": exam_address,
        "unifiedExamAddress": exam_address == "统一考试地址",
        "preLoginPrompt": get_field("考前等待提示", "考前提示", "考前等待"),
        "welcomeText": get_field("欢迎语"),
        "pledgeContent": get_field("考试承诺书内容", "承诺书内容"),
        "videoMonitor": parse_bool(get_field("视频监控")),
        "videoRecord": parse_bool(get_field("视频录制")),
        "loginVerifyMode": "考后公安验证",
        "hawkeye": parse_bool(get_field("鹰眼监控")),
        "examType": exam_type,
        "clientExam": exam_type == "客户端考试",
        "webExam": exam_type == "网页考试",
        "leaveLimit": parse_minutes(get_field("允许离开次数", "离开次数", "只允许离开次数")),
        "clientLoginLimit": parse_minutes(get_field("登陆次数", "登录次数", "允许登录次数")) or 10,
        "manualScore": parse_enabled_text(manual_score_text),
        "manualScoreText": manual_score_text,
        "watermark": parse_bool(field_map.get("答题水印")),
        "disableCopy": parse_bool(field_map.get("禁止复制")),
        "subjects": subjects,
        "courses": courses,
        "subjectImportPath": subject_import_path,
        "mockExamEnabled": bool(mock_start_dt and mock_end_dt),
        "mockExamName": f'{get_field("考试名称", "考试名")}-试考' if get_field("考试名称", "考试名") else "",
        "mockStartTimeDisplay": format_datetime(mock_start_dt),
        "mockEndTimeDisplay": format_datetime(mock_end_dt),
        "mockStartTimeIso": mock_start_dt.isoformat() if mock_start_dt else "",
        "mockEndTimeIso": mock_end_dt.isoformat() if mock_end_dt else "",
        "visibleFields": ["姓名", "身份证号"],
        "editableFields": [],
        "requiredFields": [],
        "confirmOnly": True,
    }

    warnings = []
    if not config["examName"]:
        warnings.append("缺少考试名称。")
    if not config["startTimeIso"] or not config["endTimeIso"]:
        warnings.append("考试日期时间无法解析，请检查需求单格式。")
    if get_field("试考日期时间", "试考时间", "试考起止时间") and not config["mockExamEnabled"]:
        warnings.append("试考日期时间无法解析，试考自动创建会跳过。")
    if not subjects:
        warnings.append("未读取到科目信息，批量导入科目步骤会跳过。")
    if subjects and not courses:
        warnings.append("科目信息缺少考试日期，无法按规则生成 code/form_codes。")

    return {
        "filename": path.name,
        "metrics": {
            "recognizedFields": len([value for value in field_map.values() if value]),
            "needsReview": 4,
            "etaMinutes": 4,
        },
        "previewRows": build_preview(config),
        "warnings": warnings,
        "config": config,
    }


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: exam_request_parser.py <xlsx-path>")
    result = parse_workbook(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
