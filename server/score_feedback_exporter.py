#!/usr/bin/env python3
import copy
import json
import sys
from datetime import date, datetime
from pathlib import Path


DATA_START_ROW = 7
HEADERS = ("姓名", "性别", "证件号码", "手机号码", "邮箱", "科目", "准考证号", "考试状态", "得分", "违纪情况")
TEXT_COLUMNS = (3, 4, 7)


def text(value):
    if value is None:
        return ""
    return str(value)


def normalized_status(value):
    raw = text(value).strip()
    if not raw:
        return "缺考"
    if raw == "已完成":
        return "参考"
    if raw == "未开考" or raw.lower() == "valid":
        return "缺考"
    return raw


def score_value(value):
    if value is None or text(value).strip() == "":
        return ""
    raw = text(value).strip()
    try:
        numeric = float(raw)
        return int(numeric) if numeric.is_integer() else numeric
    except ValueError:
        return raw


def find_footer_row(sheet):
    for row in range(DATA_START_ROW, sheet.max_row + 1):
        values = [text(sheet.cell(row, column).value) for column in range(1, sheet.max_column + 1)]
        joined = "".join(values)
        if "全美在线" in joined or "盖成绩" in joined:
            return row
    return sheet.max_row + 1


def copy_row_style(sheet, source_row, target_row):
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.font:
            target.font = copy.copy(source.font)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)


def clear_row_values(sheet, row):
    for column in range(1, 11):
        sheet.cell(row, column).value = None


def write_row(sheet, row_index, row, default_course):
    exam_status = normalized_status(row.get("exam_status"))
    score = score_value(row.get("score"))
    if exam_status == "缺考" and score == "":
        score = "--"
    values = [
        text(row.get("name")),
        text(row.get("gender")),
        text(row.get("identity_id")),
        text(row.get("mobile")),
        text(row.get("email")),
        text(row.get("course") or default_course),
        text(row.get("permit")),
        exam_status,
        score,
        text(row.get("violation") or "无"),
    ]
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row_index, column)
        cell.value = value
        if column in TEXT_COLUMNS:
            cell.number_format = "@"


def export_score_feedback(template_path, payload_path, output_path):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("当前 Python 环境缺少 openpyxl，请先安装依赖") from exc

    template_path = Path(template_path)
    payload_path = Path(payload_path)
    output_path = Path(output_path)
    if not template_path.exists():
        return {"ok": False, "errors": [f"成绩单模板不存在：{template_path}"]}

    payload = json.loads(payload_path.read_text("utf-8"))
    rows = payload.get("rows") or []
    workbook = load_workbook(template_path)
    sheet = workbook.active

    exam_name = text(payload.get("examName") or "")
    exam_time = text(payload.get("examTime") or "")
    today = date.today()
    processed_date = text(payload.get("processedDate") or f"{today.year}年{today.month}月{today.day}日")
    sheet["B4"] = exam_name
    sheet["C4"] = exam_time

    footer_row = find_footer_row(sheet)
    existing_slots = max(0, footer_row - DATA_START_ROW)
    required_rows = max(1, len(rows))
    if required_rows > existing_slots:
        sheet.insert_rows(footer_row, required_rows - existing_slots)
        footer_row += required_rows - existing_slots

    footer_date_cell = sheet.cell(footer_row + 1, 6)
    footer_date_cell.value = processed_date
    footer_date_cell.number_format = "@"

    last_data_row = footer_row - 1
    for row_index in range(DATA_START_ROW, last_data_row + 1):
        copy_row_style(sheet, DATA_START_ROW, row_index)
        clear_row_values(sheet, row_index)

    default_course = exam_name
    for offset, row in enumerate(rows):
        write_row(sheet, DATA_START_ROW + offset, row, default_course)

    for column in TEXT_COLUMNS:
        for row_index in range(DATA_START_ROW, max(DATA_START_ROW, DATA_START_ROW + len(rows))):
            sheet.cell(row_index, column).number_format = "@"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "ok": True,
        "path": str(output_path),
        "rows": len(rows),
        "errors": [],
    }


def main():
    if len(sys.argv) != 4:
        print(json.dumps({"ok": False, "errors": ["用法：score_feedback_exporter.py template.xlsx payload.json output.xlsx"]}, ensure_ascii=False))
        return 1
    try:
        result = export_score_feedback(sys.argv[1], sys.argv[2], sys.argv[3])
        print(json.dumps(result, ensure_ascii=False))
        return 0 if result.get("ok") else 2
    except Exception as exc:
        print(json.dumps({"ok": False, "errors": [str(exc)]}, ensure_ascii=False))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
