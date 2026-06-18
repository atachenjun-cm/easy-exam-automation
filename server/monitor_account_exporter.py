#!/usr/bin/env python3
import json
import sys
from pathlib import Path


HEADERS = ("场次ID", "场次名称", "班级名称", "考生人数", "监控账号", "监控口令", "监考地址")


def text(value):
    if value is None:
        return ""
    return str(value)


def export_monitor_accounts(payload_path, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception as exc:
        raise RuntimeError("当前 Python 环境缺少 openpyxl，请先安装依赖") from exc

    payload = json.loads(Path(payload_path).read_text("utf-8"))
    session = payload.get("session") or {}
    rooms = payload.get("rooms") or []
    if not text(session.get("session_id")).strip():
        return {"ok": False, "errors": ["缺少 session_id"]}
    if not rooms:
        return {"ok": False, "errors": ["缺少监考账号数据"]}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "监考账号"
    sheet.append(list(HEADERS))

    for room in rooms:
        sheet.append(
            [
                text(session.get("session_id")),
                text(session.get("name")),
                text(room.get("name")),
                text(room.get("num")),
                text(room.get("account")),
                text(room.get("pwd")),
                text(room.get("monitor_url") or room.get("monitorUrl") or room.get("url") or session.get("url")),
            ]
        )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows():
        for cell in row:
            cell.number_format = "@"
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = {
        "A": 14,
        "B": 32,
        "C": 16,
        "D": 12,
        "E": 16,
        "F": 16,
        "G": 52,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    workbook.save(output_path)
    return {"ok": True, "path": str(output_path), "rows": len(rooms), "errors": []}


def main():
    if len(sys.argv) != 3:
        print(json.dumps({"ok": False, "errors": ["用法：monitor_account_exporter.py payload.json output.xlsx"]}, ensure_ascii=False))
        return 1
    try:
        result = export_monitor_accounts(sys.argv[1], sys.argv[2])
        print(json.dumps(result, ensure_ascii=False))
        return 0 if result.get("ok") else 2
    except Exception as exc:
        print(json.dumps({"ok": False, "errors": [str(exc)]}, ensure_ascii=False))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
